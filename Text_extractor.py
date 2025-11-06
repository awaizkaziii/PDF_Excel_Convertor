import io
import os
import tempfile
from pathlib import Path
from typing import Dict, List, Tuple

import streamlit as st
import openpyxl
from PIL import Image
import pytesseract
import subprocess

def app():
    main()


# ==============================
# Core helpers
# ==============================

def find_tesseract_path():
    """
    Try to find Tesseract executable in common locations or system PATH
    """
    # Common installation paths
    common_paths = [
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        "/usr/bin/tesseract",
        "/usr/local/bin/tesseract",
        "/opt/homebrew/bin/tesseract"
    ]
    
    # First check if tesseract is in PATH
    try:
        subprocess.run(["tesseract", "--version"], capture_output=True, check=True)
        return "tesseract"  # Use from PATH
    except (subprocess.CalledProcessError, FileNotFoundError):
        pass
    
    # Check common installation paths
    for path in common_paths:
        if os.path.exists(path):
            return path
    
    return None


def check_tesseract_installation():
    """
    Check if Tesseract is available and provide installation instructions if not
    """
    tesseract_path = find_tesseract_path()
    
    if tesseract_path is None:
        st.error("Tesseract OCR not found. Please install Tesseract:")
        st.code("""
# Windows: Download from https://github.com/UB-Mannheim/tesseract/wiki
# macOS: brew install tesseract
# Linux (Ubuntu/Debian): sudo apt-get install tesseract-ocr
# Linux (CentOS/RHEL): sudo yum install tesseract
""")
        return False
    
    # Set the tesseract command
    if tesseract_path != "tesseract":
        pytesseract.pytesseract.tesseract_cmd = tesseract_path
    
    # Verify it works
    try:
        if tesseract_path == "tesseract":
            subprocess.run(["tesseract", "--version"], capture_output=True, check=True)
        else:
            subprocess.run([tesseract_path, "--version"], capture_output=True, check=True)
        return True
    except (subprocess.CalledProcessError, FileNotFoundError):
        st.error(f"Tesseract found at {tesseract_path} but failed to execute.")
        return False


def extract_images_grouped_by_cell(
    excel_path: str,
    output_folder: str,
    naming_convention: str = "cell"
) -> Tuple[int, Dict[Tuple[str, str], List[Dict]]]:
    """
    Extract images from the Excel workbook and group them by (sheet_name, cell_address).
    Returns:
        image_count: total number of images found
        grouped: dict keyed by (sheet, cell) -> list of image records
                 [{filename, sheet, cell, dimensions, index, per_cell_index}]
    """
    os.makedirs(output_folder, exist_ok=True)
    wb = openpyxl.load_workbook(excel_path)
    image_count = 0
    grouped: Dict[Tuple[str, str], List[Dict]] = {}

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        if hasattr(sheet, "_images") and sheet._images:
            per_cell_counter: Dict[Tuple[str, str], int] = {}
            for i, image in enumerate(sheet._images):
                try:
                    # Read image bytes (best-effort across openpyxl variations)
                    try:
                        data = image._data()
                    except Exception:
                        data = getattr(getattr(image, "image", image), "_data", lambda: b"")()

                    img = Image.open(io.BytesIO(data))

                    # Anchor -> cell address
                    anchor_from = getattr(image.anchor, "_from", None)
                    if anchor_from:
                        row = anchor_from.row + 1
                        col = anchor_from.col + 1
                        col_letter = openpyxl.utils.get_column_letter(col)
                        cell_addr = f"{col_letter}{row}"
                    else:
                        cell_addr = "Unknown"

                    key = (sheet_name, cell_addr)
                    per_cell_counter[key] = per_cell_counter.get(key, 0) + 1
                    per_idx = per_cell_counter[key]

                    # Filename
                    if naming_convention == "cell" and cell_addr != "Unknown":
                        base = f"{sheet_name}_{cell_addr}_{per_idx:02d}.png"
                    elif naming_convention == "sequential":
                        base = f"image_{image_count+1:03d}.png"
                    else:
                        base = f"{sheet_name}_{i+1:03d}.png"

                    filename = "".join(c for c in base if c.isalnum() or c in ("_", "-", "."))
                    out_path = os.path.join(output_folder, filename)
                    img.save(out_path)

                    grouped.setdefault(key, []).append({
                        "filename": filename,
                        "sheet": sheet_name,
                        "cell": cell_addr,
                        "dimensions": f"{img.size[0]}x{img.size[1]}",
                        "index": image_count,         # global extraction order
                        "per_cell_index": per_idx,    # order within the cell
                    })
                    image_count += 1

                except Exception as e:
                    print(f"[extract] {sheet_name}: {e}")

    return image_count, grouped


def ocr_replace_and_remove_images(
    excel_path: str,
    target_columns: Dict[str, List[str]],  # Changed: dict of sheet_name -> list of columns
    output_folder: str,
    tesseract_path: str = None,
    join_delimiter: str = "\n"
) -> Tuple[bool, str, int, int]:
    """
    For each cell in the target columns:
      - OCR ALL images anchored to that cell
      - Replace the cell value with the combined OCR text
      - Remove the images from that cell (so only text remains)
    Saves to *_OCR_Processed.xlsx next to the input file.
    Returns: (success, output_excel_path, total_images_found, cells_updated)
    """
    try:
        # Set tesseract path if provided, otherwise use auto-detection
        if tesseract_path and tesseract_path != "tesseract" and os.path.exists(tesseract_path):
            pytesseract.pytesseract.tesseract_cmd = tesseract_path

        # 1) Extract
        image_count, grouped = extract_images_grouped_by_cell(
            excel_path, output_folder, naming_convention="cell"
        )

        # 2) Load workbook
        wb = openpyxl.load_workbook(excel_path)
        sheets = {name: wb[name] for name in wb.sheetnames}
        updated_cells = 0

        # Helper to remove images from a specific cell
        def remove_images_from_cell(sheet_obj, cell_addr: str):
            if not (hasattr(sheet_obj, "_images") and sheet_obj._images):
                return
            to_remove = []
            for img in sheet_obj._images:
                anchor_from = getattr(img.anchor, "_from", None)
                if anchor_from:
                    r = anchor_from.row + 1
                    c = anchor_from.col + 1
                    col_letter = openpyxl.utils.get_column_letter(c)
                    addr = f"{col_letter}{r}"
                else:
                    addr = "Unknown"
                if addr == cell_addr:
                    to_remove.append(img)
            for img in to_remove:
                try:
                    sheet_obj._images.remove(img)
                except Exception:
                    pass

        # 3) OCR + replace + remove
        for (sheet_name, cell_addr), images in grouped.items():
            if cell_addr == "Unknown":
                continue
            col_letter = "".join(filter(str.isalpha, cell_addr)).upper()
            
            # Check if this cell's column is in target columns for this sheet
            if sheet_name not in target_columns:
                continue
                
            target_cols_for_sheet = {c.strip().upper() for c in target_columns[sheet_name]}
            if col_letter not in target_cols_for_sheet:
                continue

            # Stable order
            images_sorted = sorted(images, key=lambda r: (r["index"], r.get("per_cell_index", 0)))

            texts = []
            for rec in images_sorted:
                img_path = os.path.join(output_folder, rec["filename"])
                if not os.path.exists(img_path):
                    continue
                try:
                    with Image.open(img_path) as pil:
                        extracted = pytesseract.image_to_string(pil).strip()
                    cleaned = " ".join(extracted.split())
                    if cleaned:
                        texts.append(cleaned)
                except Exception as e:
                    print(f"[ocr] {sheet_name}!{cell_addr}: {e}")

            sheet = sheets[sheet_name]
            # Replace value even if OCR is empty? Here we remove images anyway to ensure text-only.
            if texts:
                sheet[cell_addr] = join_delimiter.join(texts)
                updated_cells += 1
            else:
                sheet[cell_addr] = ""  # ensure it's text-only (empty)

            remove_images_from_cell(sheet, cell_addr)

        # 4) Save output
        original = Path(excel_path)
        out_path = original.parent / f"{original.stem}_OCR_Processed.xlsx"
        wb.save(out_path)

        return True, str(out_path), image_count, updated_cells

    except Exception as e:
        return False, f"Error: {e}", 0, 0


def get_sheet_names(excel_path: str) -> List[str]:
    """Get all sheet names from the Excel file."""
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names
    except Exception as e:
        st.error(f"Error reading Excel file: {e}")
        return []

def main():
# ==============================
# Streamlit UI
# ==============================

#st.set_page_config(page_title="Upload Excel → OCR images → Text-only", page_icon="🧠", layout="centered")
    st.title("Text Extractor for Excel Images")

    with st.sidebar:
        st.header("Settings")

        # Check Tesseract installation first
        tesseract_available = check_tesseract_installation()
        
        # Only show manual path input if automatic detection failed
        tess_path = None
        if not tesseract_available:
            st.warning("Tesseract not found automatically. Please specify the path manually:")
            tess_path = st.text_input(
                "Tesseract executable path",
                value=r"C:\Program Files\Tesseract-OCR\tesseract.exe",
                help="Windows example: C:\\Program Files\\Tesseract-OCR\\tesseract.exe"
            )
        else:
            st.success("✓ Tesseract found automatically!")
            # Use the auto-detected path
            tess_path = find_tesseract_path()

        column_option = st.radio(
            "Column specification method",
            ["Same columns for all sheets", "Different columns for each sheet"],
            index=0,
            help="Choose how to specify target columns for OCR"
        )

        if column_option == "Same columns for all sheets":
            cols_str = st.text_input(
                "Target columns (comma-separated)",
                value="X",
                help="OCR will only run for images anchored in these columns for ALL sheets."
            )
            target_columns_all = [c.strip().upper() for c in cols_str.split(",") if c.strip()]
        else:
            st.info("Upload Excel file first to see sheet-specific column inputs")

        join_style = st.selectbox(
            "Join text for multiple images in the same cell with",
            ["New line", "Space", "Custom"],
            index=0
        )
        if join_style == "New line":
            join_delim = "\n"
        elif join_style == "Space":
            join_delim = " "
        else:
            join_delim = st.text_input("Custom delimiter", value="\n---\n")

    st.write("Upload an **.xlsx** file. The app will OCR images in your selected columns and remove them, leaving text only.")

    uploaded = st.file_uploader("Upload Excel (.xlsx)", type=["xlsx"])

    # If different columns per sheet is selected and file is uploaded, show sheet-specific inputs
    sheet_columns = {}
    if uploaded and column_option == "Different columns for each sheet":
        with tempfile.TemporaryDirectory() as tmpdir:
            src_path = os.path.join(tmpdir, uploaded.name)
            with open(src_path, "wb") as f:
                f.write(uploaded.getbuffer())
            sheet_names = get_sheet_names(src_path)
            
        if sheet_names:
            st.subheader("Specify columns for each sheet")
            for sheet_name in sheet_names:
                sheet_columns[sheet_name] = st.text_input(
                    f"Columns for sheet: '{sheet_name}'",
                    value="X",
                    help=f"Comma-separated columns for sheet '{sheet_name}'"
                )

    if uploaded:
        # Check if Tesseract is available before processing
        if not tesseract_available and (not tess_path or not os.path.exists(tess_path)):
            st.error("Tesseract is required but not found. Please install Tesseract or provide the correct path.")
            st.stop()

        # Prepare target_columns dictionary based on selected option
        if column_option == "Same columns for all sheets":
            with tempfile.TemporaryDirectory() as tmpdir:
                src_path = os.path.join(tmpdir, uploaded.name)
                with open(src_path, "wb") as f:
                    f.write(uploaded.getbuffer())
                sheet_names = get_sheet_names(src_path)
            
            if sheet_names:
                target_columns = {sheet_name: target_columns_all for sheet_name in sheet_names}
            else:
                target_columns = {}
        else:
            # Different columns per sheet
            target_columns = {}
            for sheet_name, cols_str in sheet_columns.items():
                if cols_str.strip():
                    target_columns[sheet_name] = [c.strip().upper() for c in cols_str.split(",") if c.strip()]

        if st.button("🚀 Run OCR and Remove Images"):
            if not target_columns:
                st.error("No target columns specified. Please provide column names.")
            else:
                with st.status("Processing…", expanded=True) as status:
                    try:
                        with tempfile.TemporaryDirectory() as tmpdir:
                            src_path = os.path.join(tmpdir, uploaded.name)
                            with open(src_path, "wb") as f:
                                f.write(uploaded.getbuffer())
                            st.write("• Saved your file to a secure temp folder")

                            images_out = os.path.join(tmpdir, "extracted_images")

                            st.write("• Extracting images, performing OCR, and removing images from target cells…")
                            ok, out_excel, total_imgs, cells_updated = ocr_replace_and_remove_images(
                                excel_path=src_path,
                                target_columns=target_columns,
                                output_folder=images_out,
                                tesseract_path=tess_path,
                                join_delimiter=join_delim
                            )

                            if not ok:
                                status.update(label="Failed", state="error")
                                st.error(out_excel if out_excel else "Unexpected error.")
                            else:
                                status.update(label="Completed", state="complete")
                                st.success(f"Done. Found {total_imgs} image(s); updated {cells_updated} cell(s).")
                                with open(out_excel, "rb") as f:
                                    st.download_button(
                                        "⬇️ Download processed Excel",
                                        f.read(),
                                        file_name=Path(out_excel).name,
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    )
                    except Exception as e:
                        status.update(label="Error", state="error")
                        st.exception(e)
    else:
        st.caption("Tip: If OCR is blank, confirm Tesseract is installed and the path above is correct.")